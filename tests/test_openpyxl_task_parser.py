from __future__ import annotations

from io import BytesIO
from importlib.util import find_spec
import sys
from pathlib import Path
import unittest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from post_bot.infrastructure.excel.openpyxl_task_parser import OpenPyxlTaskParser  # noqa: E402
from post_bot.shared.errors import ValidationError  # noqa: E402


@unittest.skipIf(find_spec("openpyxl") is None, "openpyxl is not installed")
class OpenPyxlTaskParserTests(unittest.TestCase):
    @staticmethod
    def _build_workbook_bytes(*, headers: list[str | None], rows: list[list[object]] | None = None) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        for column_index, header in enumerate(headers, start=1):
            ws.cell(row=1, column=column_index).value = header

        if rows:
            for row_index, row_values in enumerate(rows, start=2):
                for column_index, value in enumerate(row_values, start=1):
                    ws.cell(row=row_index, column=column_index).value = value

        # Simulate formatting artifacts expanding sheet width beyond contract columns.
        ws.merge_cells("O1:Y1")

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_parses_current_layout_without_legacy_columns(self) -> None:
        parser = OpenPyxlTaskParser()
        payload = self._build_workbook_bytes(
            headers=[
                "channel",
                "topic",
                "keywords",
                "response_language",
                "mode",
                "title",
                "include_image",
                "footer_text",
                "footer_link",
                "schedule_at",
            ],
            rows=[[ 
                "@news",
                "AI topic",
                "ai",
                "en",
                "instant",
                "Title",
                "FALSE",
                "Footer",
                "https://example.com",
                "2026-04-09 10:00",
            ]],
        )

        parsed = parser.parse(payload)

        self.assertEqual(len(parsed.headers), 10)
        self.assertEqual(parsed.headers[0], "channel")
        self.assertEqual(parsed.headers[-1], "schedule_at")
        self.assertEqual(len(parsed.rows), 1)
        self.assertEqual(parsed.rows[0].values["mode"], "instant")
        self.assertEqual(parsed.rows[0].values["footer_link"], "https://example.com")

    def test_legacy_removed_columns_are_parsed_and_can_be_ignored_later(self) -> None:
        parser = OpenPyxlTaskParser()
        payload = self._build_workbook_bytes(
            headers=[
                "channel",
                "topic",
                "keywords",
                "time_range",
                "response_language",
                "mode",
                "title",
                "style",
                "length",
                "include_image",
                "footer_text",
                "footer_link",
                "schedule_at",
            ],
            rows=[[ 
                "@news",
                "AI topic",
                "ai",
                "24h",
                "en",
                "instant",
                "Title",
                "journalistic",
                "medium",
                "FALSE",
                "Footer",
                "https://example.com",
                "2026-04-09 10:00",
            ]],
        )

        parsed = parser.parse(payload)

        self.assertIn("time_range", parsed.headers)
        self.assertIn("style", parsed.headers)
        self.assertIn("length", parsed.headers)
        self.assertEqual(parsed.rows[0].values["mode"], "instant")
        self.assertEqual(parsed.rows[0].values["footer_text"], "Footer")

    def test_stops_reading_after_first_empty_row(self) -> None:
        parser = OpenPyxlTaskParser()
        payload = self._build_workbook_bytes(
            headers=[
                "channel",
                "topic",
                "keywords",
                "response_language",
                "mode",
                "title",
                "include_image",
                "footer_text",
                "footer_link",
                "schedule_at",
            ],
            rows=[
                ["@news", "Row 1", "ai", "en", "instant"],
                [],
                ["@news", "Row 3", "ai", "en", "instant"],
            ],
        )

        parsed = parser.parse(payload)

        self.assertEqual(len(parsed.rows), 1)
        self.assertEqual(parsed.rows[0].excel_row, 2)
        self.assertEqual(parsed.rows[0].values["topic"], "Row 1")


    def test_supports_legacy_layout_with_search_language_column(self) -> None:
        parser = OpenPyxlTaskParser()
        payload = self._build_workbook_bytes(
            headers=[
                "channel",
                "topic",
                "title",
                "keywords",
                "search_language",
                "response_language",
                "include_image",
                "footer_text",
                "footer_link",
                "schedule_at",
                "mode",
            ],
            rows=[[ 
                "@news",
                "AI topic",
                "Title",
                "ai",
                "ar,es,en",
                "en",
                "FALSE",
                "Footer",
                "https://example.com",
                "2026-04-09 10:00",
                "instant",
            ]],
        )

        parsed = parser.parse(payload)

        self.assertEqual(len(parsed.headers), 14)
        self.assertIn("search_language", parsed.headers)
        self.assertEqual(parsed.rows[0].values["schedule_at"], "2026-04-09 10:00")
        self.assertEqual(parsed.rows[0].values["mode"], "instant")
    def test_empty_header_cell_reports_exact_excel_cell(self) -> None:
        parser = OpenPyxlTaskParser()
        payload = self._build_workbook_bytes(
            headers=["channel", None, "keywords", "response_language", "mode"],
            rows=[[ 
                "@news",
                "AI topic",
                "ai",
                "en",
                "instant",
            ]],
        )

        with self.assertRaises(ValidationError) as context:
            parser.parse(payload)

        self.assertEqual(context.exception.code, "EXCEL_HEADER_EMPTY")
        self.assertEqual(context.exception.details.get("empty_columns"), [2])
        self.assertEqual(context.exception.details.get("empty_cells"), ["B1"])


if __name__ == "__main__":
    unittest.main()

